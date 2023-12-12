import openpyxl
from datetime import datetime

class SaveToXlsx:
    def __init__(self, data_dict):
        self.data_dict = data_dict
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.output_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    def get_output_filename(self):
        return self.output_filename

    def export_to_excel(self):
        self.write_headers()
        self.write_data()
        self.workbook.save(self.get_output_filename())

    def write_headers(self):
        headers = list(self.data_dict.values())[0].keys()
        self.sheet.cell(row=1, column=1, value="link")
        for col_num, header in enumerate(headers, start=2):
            cell = self.sheet.cell(row=1, column=col_num)
            cell.value = header

    def write_data(self):
        for row_num, (link, values) in enumerate(self.data_dict.items(), start=2):
            # Zapisz link w pierwszej kolumnie
            cell = self.sheet.cell(row=row_num, column=1)
            cell.value = link

            # Zapisz resztę danych zgodnie z pierwotnym kodem
            for col_num, value in enumerate(values.values(), start=2):
                cell = self.sheet.cell(row=row_num, column=col_num)
                cell.value = value
